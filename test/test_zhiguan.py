# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# ======================================================================
#   Copyright (C) 2020 liaozhimingandy@qq.com Ltd. All Rights Reserved.
#
#   @Author      : zhiming
#   @Project     : zhiming
#   @File Name   : test_zhiguan.py	
#   @Created Date: 2020-04-16 9:04
#      @Software : PyCharm
#         @e-Mail: liaozhimingandy@qq.com
#   @Description :
#
# ======================================================================


def main():
    file_name = 'test_data.txt'
    with open(file=file_name, encoding='utf-8') as f:
        mcontent = f.readline()
        while mcontent:
            if mcontent.startswith('[填报指标]'):
                print(mcontent.replace('\n', ''))
            mcontent = f.readline()


def write2xlsx():
    file_name = 'tmp.txt'
    with open(file=file_name, encoding='utf-8') as f:
        content = f.readline()
        lst_data = list()
        while content:
            lst_data.append(content)
            content = f.readline()

    import openpyxl as xw
    file_to = xw.Workbook()
    st = file_to.create_sheet(r'处理结果', 0)

    for mcontent in lst_data:
        st.cell(lst_data.index(mcontent)+1, 1, mcontent)

    file_to.save('tmp_result.xlsx')


if __name__ == "__main__":
    # main()
    write2xlsx()