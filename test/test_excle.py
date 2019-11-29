# !/usr/bin/env python3
# -*- coding:utf-8 -*-
# 广医五院  志明  2019-05-27 18:29
# 当前计算机登录名称 :andy
# 项目名称  :andy
# 编译器   :PyCharm
__author____ = '志明'
__time__ = '2019-05-27 18:29'

import os
import random
import time

# excel文件写入工具包
import openpyxl as xt
# excel文件读取操作工具包
import xlrd as xd


def main1():
    import math
    import os
    file = r'c:\Users\andy\Desktop\test.xlsx'
    data = list()
    # 判断文件是否存在,不存在则直接跳出
    if not os.path.exists(file):
        print(r'不存在需要处理的文件<{0}>'.format(file))
        exit(1)
    f = xd.open_workbook(file).sheet_by_index(0)
    print(f.nrows, f.ncols)
    # 2,数据提取到列表中
    for row in range(math.floor(f.nrows / 2)):
        # 获取表格第一列和第二列里面值并且格式处理
        r1, r2 = str(f.row(2 * row)[0].value).strip(), f.row(2 * row + 1)[1].value.strip().replace(' ', '')
        print(row, r1, r2)
        data.append((r1, r2))
    del row
    print('整理后的数据:{0}'.format(data))
    # 分割文件路径和文件后缀名,得到新文件的路径
    file_path, file_ext = os.path.splitext(file)
    file_new = os.path.join(file_path + '-result' + file_ext)

    # 3,excel文件写入操作
    file_to = xt.Workbook()
    st = file_to.create_sheet('test')
    # st.cell(1, 1, '测试')
    i = 0
    for row in data:
        i += 1
        st.cell(i, 1, row[0])
        st.cell(i, 2, row[1])
        print('第{0}行数据写入到{1}文件成功;内容为:{2}-{3}'.format(i, file_new, row[0], row[1]))
    file_to.save(file_new)


def trans():
    """
    将excel中英文翻译成中文
    :return: None
    """
    import time
    import random
    file = r'c:\Users\andy\Desktop\test.xlsx'
    data = list()
    # 1, 判断文件是否存在,不存在则直接跳出
    if not os.path.exists(file):
        print('温馨提示:\n不存在需要处理的文件:{0}'.format(file))
        exit(1)
    f = xd.open_workbook(file).sheet_by_index(0)

    # 2,数据提取到列表中
    for row in range(f.nrows):
        # 获取表格第一列和第二列里面值并且格式处理
        r1 = str(f.row(row)[0].value).strip()
        print(row, r1)
        data.append(r1)
    del row

    print('整理后的数据:{0}'.format(data))
    # 分割文件路径和文件后缀名,得到新文件的路径
    file_path, file_ext = os.path.splitext(file)
    file_new = os.path.join(file_path + '-result' + file_ext)
    # 3,excel文件写入操作
    file_to = xt.Workbook()
    st = file_to.create_sheet('test')
    i = 0
    # 百度翻译接口调用
    from SuTranslate.BaiDuTranslator import BaiDuTranslator
    bd = BaiDuTranslator()
    for row in data:
        i += 1
        st.cell(i, 1, row)
        print(row)
        text = bd.translate(row).get('result')
        if text is None:
            print('接口被封,暂时退出!')
            exit(1)
        print(text)
        st.cell(i, 2, text)
        print(f'第{i}行数据{row}翻译成功;\n翻译结果为:{text}')
        t1 = random.random() * 2
        print(f'随机休息{t1},防止翻译接口被封!')
        time.sleep(t1)
    del i
    file_to.save(file_new)
    print(f'写入{file_new}文件成功!')
    del file_new, bd, data, row
    return None


# 将xlsx文件英文翻译成中文对象
class Trans(object):

    def __init__(self, name):
        self.name = name

    # 翻译
    def trans(self, file_path=None, transtor=None):
        # 判断文件是否存在,不存在则直接跳出
        if not os.path.exists(file_path):
            print('温馨提示:\n不存在需要处理的文件:{0}'.format(file))
            return
        f = xd.open_workbook(file).sheet_by_index(0)
        datas = list()
        # 2,数据提取到列表中
        for i in range(f.nrows):
            value = f.row(i)[0].value
            try:
                result = transtor.translate(value).get('result')
            except(TypeError,):
                print(f'翻译接口被锁,请稍后再试!')
                break
            else:
                if result is None:
                    print(f'翻译接口被锁,请稍后再试!')
                    break
                time_sleep = random.random() * 2
                print(f'行数:{i + 1},数据:{value},翻译结果:{result},随机休息:{time_sleep}秒,防止接口被封!')
                datas.append((value, result))
                time.sleep(time_sleep)
        if len(datas) == 0:
            return
        # 分割文件路径和文件后缀名,得到新文件的路径
        file_path, file_ext = os.path.splitext(file_path)
        file_new = os.path.join(file_path + '-result' + file_ext)
        # 3,excel文件写入操作
        file_to = xt.Workbook()
        st = file_to.create_sheet(f'翻译结果')
        i = 0
        for data in datas:
            i += 1
            st.cell(i, 1, data[0])
            st.cell(i, 2, data[1])

        # 报错文件并删除变量
        file_to.save(file_new)
        print(f'翻译完成,文件保存完成,路径请查看:{file_new}')
        del i, datas, file_to, f, file_ext, file_new
        return


if __name__ == '__main__':
    # 百度翻译接口调用
    from SuTranslate.BaiDuTranslator import BaiDuTranslator
    tran = Trans('demo')
    file = r'c:\Users\andy\Desktop\test.xlsx'
    BD = BaiDuTranslator()
    tran.trans(file, transtor=BD)
