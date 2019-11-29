# !/usr/bin/env python3
# -*- coding:utf-8 -*-
# 广医五院  志明  2019-11-13 14:10
# 当前计算机登录名称 :andy
# 项目名称  :andy
# 编译器   :PyCharm

# 导出所有elasticsearch分类信息
import json

import chardet  # 判断数据编码使用
# excel文件写入工具包
import openpyxl as xt
import requests


class EsToolor(object):
    def __init__(self, ip="127.0.0.1", port=9200):
        self.ip = ip
        self.port = port

    def get_data(self, type="mdms.entity.masterdatamanage.master_definition",
                 param=None):
        url = r"http://" + self.ip + r":" + str(self.port) + r"/" + type + r"/_search?size=10000"
        response = requests.get(url)
        content = response.content
        encoding = chardet.detect(content)
        datas = content.decode(encoding.get("encoding"))
        data = json.loads(datas)
        data_inner = data.get("hits").get("hits")
        return {"msg": "处理成功", "result": data_inner, "error": 0}

    def save(self, datas=None, file_path="result.xlsx"):
        # 构建xlsx保存对象
        file_to = xt.Workbook()
        st = file_to.create_sheet('result', 0)
        print(type(datas))
        for val in datas:
            # st.cell(datas.index(val),)
            if not datas.index(val):
                st.cell(1, 1, "标准分类")
                st.cell(1, 2, "标准数据编码")
                st.cell(1, 3, "标准数据名称")
                st.cell(1, 4, "分类")
            print(f"正在处理第{datas.index(val) + 1}条数据...")
            st.cell(datas.index(val) + 2, 1, val["_source"]["MASTER_DEF_NAME"])
            st.cell(datas.index(val) + 2, 2, val["_source"]["MASTER_DEF_CODE"])
            st.cell(datas.index(val) + 2, 3, val["_source"]["MASTER_DIR_NAME"])
            st.cell(datas.index(val) + 2, 4, val["_source"]["DATA_TYPE"])
            # print(val)标准分类	标准数据编码	标准数据名称	分类
        file_to.save(file_path)
        del file_to
        return {"msg": "处理成功", "result": None, "error": 0}


if __name__ == "__main__":
    ip, port = "localhost", 9200
    es = EsToolor(ip, port)
    data = es.get_data().get("result")
    data = sorted(data, key=lambda val: val.get("_source").get("DATA_TYPE")[0], reverse=False)
    es.save(data, file_path="mdm_data.xlsx")
