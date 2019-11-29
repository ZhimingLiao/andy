# !/usr/bin/env python3
# -*- coding:utf-8 -*-
# 广医五院  志明  2019-11-11 16:56
# 当前计算机登录名称 :andy
# 项目名称  :andy
# 编译器   :PyCharm
import json

# excel文件写入工具包
import openpyxl as xt

if __name__ == "__main__":
    file_path = "data.json"
    file_path_save = "data.xlsx"
    with open(file_path, encoding="utf8") as ftmp:
        content = ftmp.read()
    content_json = json.loads(content)
    # 构建xlsx保存对象
    file_to = xt.Workbook()
    st = file_to.create_sheet('result')

    for val in content_json:
        # st.cell(content_json.index(val),)
        if not content_json.index(val):
            st.cell(1, 1, "标准分类")
            st.cell(1, 2, "标准数据编码")
            st.cell(1, 3, "标准数据名称")
            st.cell(1, 4, "分类")
        print(f"正在处理第{content_json.index(val) + 1}条数据...")
        st.cell(content_json.index(val) + 2, 1, val["_source"]["MASTER_DEF_NAME"])
        st.cell(content_json.index(val) + 2, 2, val["_source"]["MASTER_DEF_CODE"])
        st.cell(content_json.index(val) + 2, 3, val["_source"]["MASTER_DIR_NAME"])
        st.cell(content_json.index(val) + 2, 4, val["_source"]["DATA_TYPE"])
        print(f"处理第{content_json.index(val) + 1}条数据成功...")
        # print(val)标准分类	标准数据编码	标准数据名称	分类
    file_to.save(file_path_save)
    del file_to
    print("保存成功")
